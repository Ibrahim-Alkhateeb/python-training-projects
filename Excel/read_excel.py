# نستخدم الباني ()Workbook لإنشاء حاوية لكل أجزاء ملف إكسل.
# يمكنك إنشاء ورقة جديدة من خلال الدالة ()create_chartsheet.
# تمكّنك الدالة ()save من حفظ ملف إكسل.

from csv import excel, excel_tab

import openpyxl
from pathlib import Path

home= Path.home()
path= Path.home() / Path('PycharmProjects','PythonProjectApps','Excel')
# excel_file= path / input('please select excel file:\n')
excel_file= path / 'Excel_file.xlsx'
open_excel_file= openpyxl.load_workbook(excel_file)

# print(open_excel_file.sheetnames)

sheet1= open_excel_file['Sheet1']
# print(sheet1.title)

sheet_active= open_excel_file.active
# print(sheet_active.title)
#
# print(sheet1['A1'].value)
# print(sheet1['B1'].value)
# print(sheet1['C1'].value)
# print(sheet1['C1'].row)
# print(sheet1['C1'].column)
# print(sheet1['C1'].coordinate)
#
# print(sheet1.cell(row=2 , column=1).value)

#print employees names
# for i in range(1,22):
#     print(sheet1.cell(row=i, column=1).value)

#print employees names and salaries and Total
total = 0
for i in range(1,sheet1.max_row+1):
    name = sheet1.cell(row=i, column=1).value
    salary = sheet1.cell(row=i, column=2).value

    if salary is not None: #insure the field is not null
        if isinstance(salary, str): #if field is str remove spaces to prepare convert to int
            salary= salary.strip()

            if salary.isdigit(): #Check if the string contains only digits
                salary = int(salary)

        if isinstance(salary, (int, float)):
            total += salary
        else:
            print(f"Warning: Invalid salary format for {name}: {salary}")
    else:
        print(f"Warning: Missing salary for {name}")  # Debugging output

    print(name , salary)

print(f'the total salary of the employees is {total}')
# print(sheet1.max_row)