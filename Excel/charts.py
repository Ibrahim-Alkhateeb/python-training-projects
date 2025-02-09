import openpyxl
from pathlib import Path
from openpyxl import chart

home= Path.home()
path= Path.home() / Path('PycharmProjects','PythonProjectApps','Excel')
# excel_file= path / input('please select excel file:\n')
excel_file= path / 'Excel_file.xlsx'
open_excel_file= openpyxl.load_workbook(excel_file)
# print(open_excel_file.sheetnames)
sheet1= open_excel_file['Sheet1']

title= openpyxl.chart.Reference(sheet1, min_col=1, max_col=1, min_row=2, max_row=sheet1.max_row+1)
data= openpyxl.chart.Reference(sheet1, min_col=2, max_col=2, min_row=2, max_row=sheet1.max_row+1)
# bar_chart= openpyxl.chart.BarChart()
bar_chart= openpyxl.chart.BarChart3D()

bar_chart.title= 'My Bar Chart'
bar_chart.add_data(data=data)
bar_chart.set_categories(title)

sheet1.add_chart(bar_chart, 'E8')

open_excel_file.save(filename= excel_file)